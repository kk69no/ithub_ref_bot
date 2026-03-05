"""Excel export functionality for referrals and payments data."""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


async def export_full_report(db) -> io.BytesIO:
    """Export full referrals and payments report to Excel.

    Args:
        db: Database connection

    Returns:
        BytesIO object containing XLSX file
    """
    workbook = Workbook()

    # Remove default sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Create sheets
    await _create_students_sheet(workbook, db)
    await _create_referrals_sheet(workbook, db)
    await _create_payments_sheet(workbook, db)

    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


async def _create_students_sheet(workbook, db):
    """Create students sheet."""
    ws = workbook.create_sheet("Студенты")

    # Headers
    headers = ["ID", "Имя", "Группа", "Email", "Статус", "Дата создания", "Рефералов"]
    ws.append(headers)

    # Style headers
    _style_headers(ws, len(headers))

    # Get data
    students = await db.fetchall(
        """
        SELECT s.user_id, s.name, s.group, s.email, s.status, s.created_at,
               COUNT(r.id) as referral_count
        FROM students s
        LEFT JOIN referrals r ON s.user_id = r.curator_id
        GROUP BY s.user_id
        ORDER BY s.name
        """
    )

    # Add data
    for student in students:
        ws.append([
            student['user_id'],
            student['name'],
            student['group'],
            student['email'],
            student['status'],
            student['created_at'],
            student['referral_count'] or 0
        ])

    # Auto-width columns
    _auto_width_columns(ws)


async def _create_referrals_sheet(workbook, db):
    """Create referrals sheet."""
    ws = workbook.create_sheet("Рефералы")

    # Headers
    headers = ["ID", "Студент", "Куратор", "Статус", "Дата создания"]
    ws.append(headers)

    # Style headers
    _style_headers(ws, len(headers))

    # Get data
    referrals = await db.fetchall(
        """
        SELECT r.id, s.name as student_name, c.name as curator_name, r.status, r.created_at
        FROM referrals r
        JOIN students s ON r.student_id = s.user_id
        JOIN students c ON r.curator_id = c.user_id
        ORDER BY r.created_at DESC
        """
    )

    # Add data
    for ref in referrals:
        ws.append([
            ref['id'],
            ref['student_name'],
            ref['curator_name'],
            ref['status'],
            ref['created_at']
        ])

    # Auto-width columns
    _auto_width_columns(ws)


async def _create_payments_sheet(workbook, db):
    """Create payments sheet."""
    ws = workbook.create_sheet("Платежи")

    # Headers
    headers = ["ID", "Студент", "Сумма", "Причина", "Статус", "Дата"]
    ws.append(headers)

    # Style headers
    _style_headers(ws, len(headers))

    # Get data
    payments = await db.fetchall(
        """
        SELECT p.id, s.name, p.amount, p.reason, p.status, p.created_at
        FROM payments p
        JOIN students s ON p.user_id = s.user_id
        ORDER BY p.created_at DESC
        """
    )

    # Add data with currency formatting
    for payment in payments:
        row = ws.max_row + 1
        ws[f"A{row}"] = payment['id']
        ws[f"B{row}"] = payment['name']
        ws[f"C{row}"] = payment['amount']
        ws[f"C{row}"].number_format = '#,##0.00 "₽"'
        ws[f"D{row}"] = payment['reason']
        ws[f"E{row}"] = payment['status']
        ws[f"F{row}"] = payment['created_at']

    # Auto-width columns
    _auto_width_columns(ws)


def _style_headers(ws, column_count):
    """Apply header styling."""
    header_fill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, column_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def _auto_width_columns(ws):
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
